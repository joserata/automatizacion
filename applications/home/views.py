from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils import timezone
from django.db.models import Count
from applications.comunicaciones.forms import (
    ReglaForm,
    ResponsableForm,
)
from applications.comunicaciones.forms_consecutivo import (
    ConsecutivoForm,
)
from applications.comunicaciones.models import (
    Comunicacion,
    Consecutivo,
    Asignacion,
    Historial,
    Regla,
    Responsable,
)
from applications.comunicaciones.services import (
    GmailSyncError,
    sincronizar_gmail,
    _servicio_gmail,
    _enviar_delegacion_gmail,
)
from applications.usuarios.decorators import grupo_requerido

from .dashboard import Dashboard


import json
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.shortcuts import redirect, render
from django.utils import timezone

from applications.comunicaciones.models import Consecutivo
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required

from django.utils import timezone

from applications.comunicaciones.forms_consecutivo import ConsecutivoForm
from applications.comunicaciones.models import Comunicacion

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
# ==========================================================
# DASHBOARD
# ==========================================================

@grupo_requerido(
    "Administrador",
    "Transaccion",
    "Operador",
    "Consultor",
)
def dashboard(request):
    return render(
        request,
        "home/dashboard.html",
        Dashboard.datos(),
    )
def es_administrador(user):
    return user.is_authenticated and user.is_superuser

@login_required
@user_passes_test(es_administrador)
def consecutivos_admin(request):

    from applications.comunicaciones.models import Consecutivo

    consecutivos = (
        Consecutivo.objects
        .select_related("usuario")
        .all()
        .order_by("-id")
    )

    return render(
        request,
        "home/consecutivos_admin.html",
        {
            "consecutivos": consecutivos,
        }
    )
# ==========================================================
# ENTRADAS
# ==========================================================

@login_required
def entrada(request):

    comunicaciones = (
        Comunicacion.objects
        .filter(tipo="ENTRADA")
        .select_related("responsable", "usuario")
        .order_by("-fecha")
    )

    # ======================================================
    # ADMINISTRADOR
    # ======================================================
    # Ve absolutamente todas las entradas.
    if request.user.is_superuser or request.user.groups.filter(
        name="Administrador"
    ).exists():

        pass

    # ======================================================
    # TRANSACCION
    # ======================================================
    elif request.user.groups.filter(name="Transaccion").exists():

        comunicaciones = comunicaciones.filter(
            remitido_transaccion=True
        )

    # ======================================================
    # OPERADOR
    # ======================================================
    elif request.user.groups.filter(name="Operador").exists():

        responsable = (
            Responsable.objects
            .filter(
                usuario=request.user,
                activo=True
            )
            .first()
        )

        if responsable:
            comunicaciones = comunicaciones.filter(
                responsable=responsable,
                estado="DELEGADO"
            )
        else:
            comunicaciones = comunicaciones.none()

    # ======================================================
    # CONSULTOR
    # ======================================================
    elif request.user.groups.filter(name="Consultor").exists():

        # El consultor puede consultar las entradas.
        pass

    # ======================================================
    # CUALQUIER OTRO USUARIO
    # ======================================================
    else:

        comunicaciones = comunicaciones.none()

    responsables = (
        Responsable.objects
        .filter(activo=True)
        .order_by("nombre")
    )

    return render(
        request,
        "home/entrada.html",
        {
            "comunicaciones": comunicaciones,
            "responsables": responsables,
        },
    )

# ==========================================================
# SALIDAS
# ==========================================================

@login_required
def salida(request):

    comunicaciones = (
        Comunicacion.objects
        .filter(tipo="SALIDA")
        .select_related("usuario")
        .order_by("-fecha")
    )

    # ======================================================
    # ADMINISTRADOR
    # ======================================================

    if request.user.is_superuser or request.user.groups.filter(
        name="Administrador"
    ).exists():

        # Administrador puede ver todas.
        pass

    # ======================================================
    # CUALQUIER OTRO USUARIO
    # ======================================================

    else:

        # Cada usuario solamente ve sus propias salidas.
        comunicaciones = comunicaciones.filter(
            usuario=request.user
        )

    return render(
        request,
        "home/salida.html",
        {
            "comunicaciones": comunicaciones,
        },
    )

    # ======================================================
    # ADMINISTRADOR
    # ======================================================
    if request.user.groups.filter(name="Administrador").exists():

        # El administrador puede ver todas.
        pass

    # ======================================================
    # DEMÁS USUARIOS
    # ======================================================
    else:

        comunicaciones = comunicaciones.filter(
            usuario=request.user
        )

    return render(
        request,
        "home/salida.html",
        {
            "comunicaciones": comunicaciones,
        },
    )


# ==========================================================
# SINCRONIZAR GMAIL
# ==========================================================

@login_required
def sincronizar_correo(request):

    if request.method != "POST":
        return redirect("home:dashboard")

    try:
        creados, actualizados = sincronizar_gmail()

    except GmailSyncError as error:

        messages.error(
            request,
            str(error),
        )

    else:

        messages.success(
            request,
            (
                f"Sincronización terminada: "
                f"{creados} nuevos y "
                f"{actualizados} actualizados."
            ),
        )

    return redirect(
        request.POST.get("next") or "home:dashboard"
    )


# ==========================================================
# REMITIR A TRANSACCIÓN
# ==========================================================

@login_required
def remitir_a_transaccion(request):

    if request.method != "POST":

        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Método no permitido.",
            },
            status=405,
        )

    remitidos = request.POST.getlist("remitidos")

    procesados = []

    for valor in remitidos:

        try:
            comunicacion_id = int(valor)

        except (TypeError, ValueError):
            continue

        comunicacion = get_object_or_404(
            Comunicacion,
            pk=comunicacion_id,
        )

        Historial.objects.create(
            comunicacion=comunicacion,
            usuario=request.user,
            accion="Remitido a Transaccion",
            descripcion="Remitido por administrador",
        )

        procesados.append(
            comunicacion.id
        )

    return JsonResponse(
        {
            "ok": True,
            "mensaje": (
                f"Se remitieron "
                f"{len(procesados)} "
                f"comunicaciones a Transacción."
            ),
        }
    )


# ==========================================================
# ASIGNAR RESPONSABLES
# ==========================================================

@login_required
@grupo_requerido("Transaccion")
def asignar_responsables(request):

    if request.method != "POST":
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Método no permitido.",
            },
            status=405,
        )

    asignaciones = request.POST.getlist("asignaciones")
    print("==========================================")
    print("ASIGNACIONES RECIBIDAS")
    print("==========================================")
    print("POST completo:", request.POST)
    print("asignaciones:", asignaciones)
    print("==========================================")

    
    procesados = []
    asignaciones_creadas = 0
    asignaciones_existentes = 0

    for valor in asignaciones:
        print("VALOR DE ASIGNACION RECIBIDO:", valor)
        if ":" not in valor:
            continue

        comunicacion_id_raw, responsable_id_raw = (
            valor.split(":", 1)
        )
        print(
    "COMUNICACION:",
    comunicacion_id_raw,
    "RESPONSABLE:",
    responsable_id_raw,
)
        try:
            comunicacion_id = int(comunicacion_id_raw)
            responsable_id = int(responsable_id_raw)

        except (TypeError, ValueError):
            continue

        comunicacion = get_object_or_404(
            Comunicacion,
            pk=comunicacion_id,
        )

        responsable = get_object_or_404(
            Responsable,
            pk=responsable_id,
        )

        # ==================================================
        # CREAR ASIGNACIÓN
        # ==================================================

        asignacion, creada = Asignacion.objects.get_or_create(
            comunicacion=comunicacion,
            responsable=responsable,
            activa=True,
            defaults={
                "usuario_asignador": request.user,
            },
        )

        if creada:
            asignaciones_creadas += 1
        else:
            asignaciones_existentes += 1

        # ==================================================
        # RESPONSABLE PRINCIPAL
        # ==================================================

        if not comunicacion.responsable_id:
            comunicacion.responsable = responsable

        comunicacion.estado = "DELEGADO"

        comunicacion.save(
            update_fields=[
                "responsable",
                "estado",
            ]
        )

        # ==================================================
        # ENVIAR CORREO DE DELEGACIÓN
        # ==================================================

        if creada:
            try:
                servicio_gmail = _servicio_gmail()

                enviado = _enviar_delegacion_gmail(
                    servicio_gmail,
                    comunicacion,
                    responsable,
                )

                if enviado:
                    print(
                    "CORREO DE DELEGACIÓN ENVIADO:",
                    responsable.correo,
                )
                else:
                    print(
                    "CORREO DE DELEGACIÓN NO ENVIADO:",
                    responsable.correo,
            )

            except Exception as error:
                print(
                "ERROR ENVIANDO CORREO DE DELEGACIÓN:",
                 error,
            )
        # ==================================================
        # HISTORIAL
        # ==================================================

        Historial.objects.create(
            comunicacion=comunicacion,
            usuario=request.user,
            accion="Delegado a responsable",
            descripcion=(
                f"Delegado a {responsable.nombre}"
            ),
        )

        procesados.append(comunicacion.id)

    return JsonResponse(
        {
            "ok": True,
            "mensaje": (
                f"Se procesaron "
                f"{len(procesados)} asignaciones."
            ),
            "asignaciones_creadas": asignaciones_creadas,
            "asignaciones_existentes": asignaciones_existentes,
        }
    )

# ==========================================================
# REPORTES
# ==========================================================

@login_required
def reportes(request):

    return render(
        request,
        "home/reportes.html",
    )


# ==========================================================
# REGLAS
# ==========================================================

@login_required
def reglas(request):

    reglas = (
        Regla.objects
        .select_related("responsable")
        .order_by("palabra")
    )

    return render(
        request,
        "home/reglas.html",
        {
            "reglas": reglas,
        },
    )


# ==========================================================
# EDITAR / CREAR REGLA
# ==========================================================

@login_required
def editar_regla(request, regla_id=None):

    regla = (
        get_object_or_404(Regla, pk=regla_id)
        if regla_id
        else None
    )

    form = ReglaForm(
        request.POST or None,
        instance=regla
    )

    if request.method == "POST" and form.is_valid():

        regla_guardada = form.save()

        messages.success(
            request,
            "Regla guardada correctamente."
        )

        return redirect("home:reglas")

    return render(
        request,
        "home/formulario.html",
        {
            "form": form,
            "titulo": "Regla de delegación",
        }
    )
# ==========================================================
# EDITAR / CREAR RESPONSABLE
# ==========================================================

@login_required
def editar_responsable(
    request,
    responsable_id=None,
):

    responsable = (
        get_object_or_404(
            Responsable,
            pk=responsable_id,
        )
        if responsable_id
        else None
    )

    form = ResponsableForm(
        request.POST or None,
        instance=responsable,
    )

    if request.method == "POST" and form.is_valid():

        responsable_guardado = form.save()

        messages.success(
            request,
            (
                f"Responsable "
                f"{responsable_guardado.nombre} "
                "guardado correctamente."
            ),
        )

        return redirect(
            "home:reglas"
        )

    return render(
        request,
        "home/formulario.html",
        {
            "form": form,
            "titulo": "Responsable",
        },
    )


# ==========================================================
# REMITIR COMUNICACIONES A TRANSACCIÓN
# ==========================================================

@require_POST
@grupo_requerido("Administrador")
def remitir_transaccion(request):

    ids = request.POST.getlist(
        "ids[]"
    )

    cantidad = (
        Comunicacion.objects
        .filter(
            id__in=ids,
            tipo="ENTRADA",
        )
        .update(
            remitido_transaccion=True,
            fecha_remision=timezone.now(),
        )
    )

    return JsonResponse(
        {
            "ok": True,
            "cantidad": cantidad,
        }
    )


# ==========================================================
# ENTRADA DE TRANSACCIÓN
# ==========================================================

@login_required
@grupo_requerido("Transaccion")
def entrada_tran(request):

    comunicaciones = (
        Comunicacion.objects
        .filter(
            tipo="ENTRADA",
            remitido_transaccion=True,
        )
        .select_related("responsable")
        .prefetch_related(
            "asignaciones__responsable"
        )
        .order_by("-fecha")
    )

    responsables = (
        Responsable.objects
        .filter(activo=True)
        .order_by("nombre")
    )

    return render(
        request,
        "home/entrada_tran.html",
        {
            "comunicaciones": comunicaciones,
            "responsables": responsables,
        },
    )


# ==========================================================
# NUEVO CONSECUTIVO
# ==========================================================

# ==========================================================
# NUEVO CONSECUTIVO
# ==========================================================

@login_required
def nuevo_consecutivo(request):

    anio = timezone.now().year

    # ======================================================
    # BUSCAR ÚLTIMO CONSECUTIVO DEL AÑO
    # ======================================================

    ultimo = (
        Consecutivo.objects
        .filter(fecha__year=anio)
        .order_by("-id")
        .first()
    )

    ultimo_numero = 0

    if ultimo and ultimo.consecutivo:

        try:
            partes = ultimo.consecutivo.split("-")

            # N.1.014-0003-26
            ultimo_numero = int(partes[1])

        except (ValueError, IndexError):
            ultimo_numero = 0

    # ======================================================
    # SIGUIENTE CONSECUTIVO
    # ======================================================

    siguiente_numero = ultimo_numero + 1

    siguiente_consecutivo = (
        f"N.1.014-"
        f"{siguiente_numero:04d}-"
        f"{str(anio)[2:]}"
    )

    # ======================================================
    # POST
    # ======================================================

    if request.method == "POST":

        form = ConsecutivoForm(request.POST)

        if form.is_valid():

            try:

                with transaction.atomic():

                    # Volvemos a consultar dentro de la
                    # transacción para evitar duplicados.

                    ultimo = (
                        Consecutivo.objects
                        .select_for_update()
                        .filter(fecha__year=anio)
                        .order_by("-id")
                        .first()
                    )

                    ultimo_numero = 0

                    if ultimo and ultimo.consecutivo:

                        try:

                            partes = ultimo.consecutivo.split("-")

                            ultimo_numero = int(partes[1])

                        except (ValueError, IndexError):

                            ultimo_numero = 0

                    # ======================================
                    # GENERAR CONSECUTIVO GLOBAL
                    # ======================================

                    siguiente_numero = ultimo_numero + 1

                    numero = (
                        f"N.1.014-"
                        f"{siguiente_numero:04d}-"
                        f"{str(anio)[2:]}"
                    )

                    # ======================================
                    # GUARDAR
                    # ======================================

                    consecutivo = form.save(commit=False)

                    consecutivo.consecutivo = numero
                    consecutivo.usuario = request.user

                    consecutivo.save()

                messages.success(
                    request,
                    f"Consecutivo {numero} generado correctamente."
                )

                return redirect("home:salida")

            except Exception as e:

                messages.error(
                    request,
                    f"No fue posible generar el consecutivo: {e}"
                )

    else:

        form = ConsecutivoForm()

    # ======================================================
    # MOSTRAR ÚLTIMO Y SIGUIENTE
    # ======================================================

    contexto = {
        "form": form,
        "titulo": "Nuevo consecutivo",
        "ultimo_consecutivo": (
            ultimo.consecutivo
            if ultimo
            else "No existen consecutivos"
        ),
        "siguiente_consecutivo": siguiente_consecutivo,
    }

    return render(
        request,
        "home/consecutivo_nuevo.html",
        contexto
    )

from django.core.exceptions import PermissionDenied    

# ==========================================================
# CONSULTA DE CONSECUTIVOS DEL USUARIO
# ==========================================================

@login_required
def consulta_consecutivos(request):

    consecutivos = (
        Consecutivo.objects
        .filter(usuario=request.user)
        .select_related("usuario")
        .order_by("-fecha_creacion", "-id")
    )

    return render(
        request,
        "home/consulta_consecutivos.html",
        {
            "consecutivos": consecutivos,
            "usuario_actual": request.user,
        },
    )


# ==========================================================
# CONSULTA DE CONSECUTIVOS - ADMINISTRADOR
# ==========================================================

@login_required
def consecutivos_admin(request):

    es_administrador = (
        request.user.is_superuser
        or request.user.groups.filter(
            name="Administrador"
        ).exists()
    )

    if not es_administrador:
        raise PermissionDenied(
            "No tiene permisos para consultar todos los consecutivos."
        )

    consecutivos = (
        Consecutivo.objects
        .select_related("usuario")
        .order_by("-fecha_creacion", "-id")
    )

    return render(
        request,
        "home/consecutivos_admin.html",
        {
            "consecutivos": consecutivos,
            "usuario_actual": request.user,
        },
    )

# ==========================================================
# CONSOLIDADO DE RADICADOS
# ==========================================================

@login_required
def consolidado_radicados(request):

    radicados = (
        Comunicacion.objects
        .filter(
            tipo="ENTRADA",
            radicado__isnull=False
        )
        .select_related("responsable")
        .order_by("-fecha")
    )

    total_radicados = radicados.count()

    total_delegados = radicados.filter(
        estado="DELEGADO"
    ).count()

    total_nuevos = radicados.filter(
        estado="NUEVO"
    ).count()

    total_en_proceso = radicados.filter(
        estado="EN_PROCESO"
    ).count()

    return render(
        request,
        "home/consolidado_radicados.html",
        {
            "radicados": radicados,
            "total_radicados": total_radicados,
            "total_delegados": total_delegados,
            "total_nuevos": total_nuevos,
            "total_en_proceso": total_en_proceso,
        }
    )    

# ==========================================================
# EXPORTAR RADICADOS A EXCEL
# ==========================================================

@login_required
def exportar_radicados_excel(request):

    radicados = (
        Comunicacion.objects
        .filter(
            tipo="ENTRADA",
            radicado__isnull=False
        )
        .select_related("responsable")
        .order_by("-fecha")
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Radicados"

    encabezados = [
        "Radicado",
        "Fecha",
        "Remitente",
        "Destinatarios",
        "Asunto",
        "Referencia",
        "Responsable",
        "Estado",
        "Prioridad",
        "Remitido Transacción",
        "Fecha Remisión",
    ]

    ws.append(encabezados)

    for celda in ws[1]:

        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for c in radicados:

        ws.append([
            c.radicado or "",
            c.fecha.strftime("%d/%m/%Y %H:%M")
            if c.fecha else "",
            c.remitente or "",
            c.destinatarios or "",
            c.asunto or "",
            c.referencia or "",
            c.responsable.nombre
            if c.responsable else "Sin asignar",
            c.get_estado_display(),
            c.get_prioridad_display(),
            "SI" if c.remitido_transaccion else "NO",
            c.fecha_remision.strftime("%d/%m/%Y %H:%M")
            if c.fecha_remision else "",
        ])

    # Ajustar columnas

    anchos = [
        22,
        20,
        35,
        35,
        55,
        25,
        30,
        20,
        15,
        22,
        22,
    ]

    for numero, ancho in enumerate(
        anchos,
        start=1
    ):

        ws.column_dimensions[
            get_column_letter(numero)
        ].width = ancho

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="radicados.xlsx"'
    )

    wb.save(response)

    return response    

# ==========================================================
# EXPORTAR CONSECUTIVOS A EXCEL
# ==========================================================

@login_required
def exportar_consecutivos_excel(request):

    consecutivos = (
        Consecutivo.objects
        .select_related("usuario")
        .order_by("id")
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Consecutivos"

    encabezados = [
        "Consecutivo",
        "Fecha",
        "Dirigido a",
        "Asunto",
        "Funcionario Responsable",
        "Caso Aranda",
        "Observaciones",
        "Fecha Envío",
        "Tipo Archivo / TRD",
        "Ubicación",
        "Archivado",
        "Usuario",
        "Fecha Creación",
    ]

    ws.append(encabezados)

    for celda in ws[1]:

        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            "solid",
            fgColor="548235"
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for c in consecutivos:

        ws.append([
            c.consecutivo,
            c.fecha.strftime("%d/%m/%Y")
            if c.fecha else "",
            c.dirigido_a or "",
            c.asunto or "",
            c.funcionario_responsable or "",
            c.caso_aranda or "",
            c.observaciones or "",
            c.fecha_envio.strftime("%d/%m/%Y")
            if c.fecha_envio else "",
            c.tipo_archivo or "",
            c.ubicacion or "",
            c.archivado or "",
            c.usuario.username
            if c.usuario else "Sin usuario",
            c.fecha_creacion.strftime(
                "%d/%m/%Y %H:%M"
            )
            if c.fecha_creacion else "",
        ])

    anchos = [
        22,
        15,
        35,
        55,
        30,
        25,
        45,
        18,
        25,
        40,
        12,
        20,
        22,
    ]

    for numero, ancho in enumerate(
        anchos,
        start=1
    ):

        ws.column_dimensions[
            get_column_letter(numero)
        ].width = ancho

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="consecutivos.xlsx"'
    )

    wb.save(response)

    return response


# ==========================================================
# TABLERO DE CONTROL - CONSULTOR---13-08-2026
# ==========================================================

@login_required
@grupo_requerido("Consultor")
def tablero_control(request):



    # ------------------------------------------------------
    # COMUNICACIONES
    # ------------------------------------------------------

    total_comunicaciones = Comunicacion.objects.count()

    total_radicados = Comunicacion.objects.exclude(
        radicado__isnull=True
    ).exclude(
        radicado=""
    ).count()

    total_sin_radicado = (
        total_comunicaciones - total_radicados
    )

    # ------------------------------------------------------
    # DELEGACIONES
    # ------------------------------------------------------

    total_delegadas = Comunicacion.objects.filter(
        estado="DELEGADO"
    ).count()

    asignaciones_activas = Asignacion.objects.filter(
        activa=True
    ).count()

    responsables_activos = Responsable.objects.count()

    # ------------------------------------------------------
    # PENDIENTES
    # ------------------------------------------------------

    total_pendientes = Comunicacion.objects.filter(
        estado__in=["NUEVO", "PENDIENTE"]
    ).count()

    # ------------------------------------------------------
    # CONSECUTIVOS
    # ------------------------------------------------------

    total_consecutivos = Comunicacion.objects.exclude(
        consecutivo__isnull=True
    ).exclude(
        consecutivo=""
    ).count()

    total_archivados = Comunicacion.objects.filter(
    es_archivado=True
    ).count()

    total_pendientes_archivo = (
        total_consecutivos - total_archivados
    )

    # ------------------------------------------------------
    # PORCENTAJE DE RADICACIÓN
    # ------------------------------------------------------

    if total_comunicaciones:
        porcentaje_radicacion = round(
            (total_radicados / total_comunicaciones) * 100,
            1
        )
    else:
        porcentaje_radicacion = 0

    # ------------------------------------------------------
    # ESTADOS
    # ------------------------------------------------------

    estados_query = (
        Comunicacion.objects
        .values("estado")
        .annotate(cantidad=Count("id"))
        .order_by("-cantidad")
    )

    estado_comunicaciones = [
        {
            "estado": item["estado"] or "Sin estado",
            "cantidad": item["cantidad"],
        }
        for item in estados_query
    ]

    # ------------------------------------------------------
    # CONTEXTO
    # ------------------------------------------------------

    contexto = {
        "total_comunicaciones": total_comunicaciones,
        "total_radicados": total_radicados,
        "total_sin_radicado": total_sin_radicado,
        "total_delegadas": total_delegadas,
        "total_pendientes": total_pendientes,

        "asignaciones_activas": asignaciones_activas,
        "responsables_activos": responsables_activos,

        "total_consecutivos": total_consecutivos,
        "total_archivados": total_archivados,
        "total_pendientes_archivo": total_pendientes_archivo,

        "porcentaje_radicacion": porcentaje_radicacion,

        

        "estado_comunicaciones": json.dumps(
            estado_comunicaciones,
            ensure_ascii=False
        ),
    }

    return render(
        request,
        "home/tablero_control.html",
        {
            "total_comunicaciones": total_comunicaciones,
            "total_radicados": total_radicados,
            "total_delegadas": total_delegadas,
            "total_pendientes": total_pendientes,
            "total_sin_radicado": total_sin_radicado,
            "porcentaje_radicacion": porcentaje_radicacion,
            "asignaciones_activas": asignaciones_activas,
            "responsables_activos": responsables_activos,
            "total_consecutivos": total_consecutivos,
            "total_archivados": total_archivados,
            "total_pendientes_archivo": total_pendientes_archivo,
            
            "estado_comunicaciones": estado_comunicaciones,
        },
    )

# ---------------------------------------------------------
# AVANCE DE PROYECTO-RPA JEAC  
# --------------------------------------------------------